"""
conftest.py - Shared fixtures and configuration for Legal Risk Analyzer E2E tests
Enhanced with:
  - Session-scoped shared user (_SESSION_AUTH) to avoid rate-limit cascades
  - Safe JSON helper (_j) for 429/empty-body responses
  - 6-sheet Excel report engine
"""
import pytest

def pytest_sessionstart(session):
    import os
    if os.path.exists("worker_results.jsonl"):
        try: os.remove("worker_results.jsonl")
        except: pass
import time
import datetime
import os
import re
import requests as _req
import uuid
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# ─── Configuration ────────────────────────────────────────────────────────────
BASE_URL     = os.environ.get("BASE_URL", "https://legal-risk-analyzer-pdd.onrender.com")
FRONTEND_URL = os.environ.get("FRONTEND_URL", "https://legal-risk-analyzer-pdd.vercel.app")

TEST_EMAIL    = "testuser_e2e@legalrisk.dev"
TEST_PASSWORD = "TestPass@123"
TEST_NAME     = "E2E Test User"
TEST_DOB      = "1995-06-15"
TEST_SECURITY = "testfriend"

# ─── Session-level shared auth (populated once, reused by all modules) ────────
_SESSION_AUTH = {
    "email":    "e2e_shared_session@legalrisk.dev",
    "password": "SharedE2E@999",
    "dob":      "1990-01-15",
    "security": "sharedfriend",
    "token":    None,
}


def _j(r, default=None):
    """Safe JSON parser — returns {} (or default) if body is empty/non-JSON."""
    try:
        if r.content and len(r.content) > 0:
            return r.json()
    except Exception:
        pass
    return {} if default is None else default


def _login(email, password, retries=3):
    """Attempt login with retry on 429. Returns token string or None."""
    for _ in range(retries):
        try:
            r = _req.post(f"{BASE_URL}/login",
                data={"username": email, "password": password},
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                timeout=25)
            if r.status_code == 200:
                return _j(r).get("access_token")
            if r.status_code == 429:
                time.sleep(32)
                continue
            break
        except Exception:
            time.sleep(5)
    return None


def _signup_and_login(name, email, password, dob, security_answer, retries=3):
    """Create user (or login if exists). Returns token string or None."""
    for attempt in range(retries):
        try:
            r = _req.post(f"{BASE_URL}/signup", json={
                "name": name, "email": email, "password": password,
                "dob": dob, "is_major": True, "security_answer": security_answer
            }, timeout=25)
            if r.status_code == 200:
                tok = _j(r).get("access_token")
                if tok:
                    return tok
            # 400 = already exists → login
            if r.status_code in (400, 200):
                tok = _login(email, password)
                if tok:
                    return tok
            if r.status_code == 429:
                time.sleep(32)
                continue
        except Exception:
            time.sleep(5)
    return _login(email, password)


# ─── Session fixture — runs ONCE before any test ──────────────────────────────
@pytest.fixture(scope="session", autouse=True)
def _create_session_user():
    """Create the shared session user once. All modules can fall back to this token."""
    e = _SESSION_AUTH["email"]
    p = _SESSION_AUTH["password"]
    tok = _signup_and_login(
        "E2E Shared Session", e, p,
        _SESSION_AUTH["dob"], _SESSION_AUTH["security"]
    )
    if not tok:
        time.sleep(35)
        tok = _login(e, p)
    _SESSION_AUTH["token"] = tok
    yield


# ─── Driver factory ───────────────────────────────────────────────────────────
def _make_driver():
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--headless=new")
    service = Service(ChromeDriverManager().install())
    drv = webdriver.Chrome(service=service, options=options)
    drv.set_page_load_timeout(30)
    drv.implicitly_wait(8)
    return drv


def _quit_driver(drv):
    try:
        drv.quit()
    except Exception:
        pass


@pytest.fixture(scope="class")
def driver():
    """Class-scoped Chrome WebDriver."""
    drv = _make_driver()
    yield drv
    _quit_driver(drv)


@pytest.fixture(scope="module")
def fresh_driver():
    drv = _make_driver()
    yield drv
    _quit_driver(drv)


# ─── Helpers ──────────────────────────────────────────────────────────────────
_TC_RE = re.compile(r"(tc\d+)", re.IGNORECASE)


def _extract_tc_id(test_name: str) -> str:
    m = _TC_RE.search(test_name)
    return m.group(1).upper() if m else ""


def _extract_desc(item) -> str:
    doc = getattr(item.function, "__doc__", "") or ""
    first = doc.strip().split("\n")[0].strip()
    return re.sub(r"^TC\d+[:.]?\s*", "", first, flags=re.IGNORECASE).strip()


def _classify_type(category: str) -> str:
    cat = category.lower()
    if any(k in cat for k in ("selenium", "login", "signup", "upload", "dashboard",
                               "chat", "translat", "history", "profile",
                               "settings", "templates", "export")):
        return "UI / Selenium"
    if any(k in cat for k in ("api", "auth", "health", "analyze")):
        return "API"
    if any(k in cat for k in ("edge", "boundary", "security", "performance", "extended")):
        return "Security / Edge"
    return "E2E"


# ─── Result store ─────────────────────────────────────────────────────────────
_test_results = []


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    if rep.when == "call" or (rep.when in ["setup", "teardown"] and rep.failed):
        error_msg = ""
        if rep.failed:
            error_msg = rep.longreprtext if hasattr(rep, "longreprtext") else str(rep.longrepr)
        category = (item.nodeid.split("::")[0]
                    .split("/")[-1].split("\\")[-1].replace(".py", ""))
        res = {
            "tc_id":       _extract_tc_id(item.name),
            "category":    category,
            "test_name":   item.name,
            "description": _extract_desc(item),
            "type":        _classify_type(category),
            "outcome":     rep.outcome,
            "duration":    getattr(rep, "duration", 0),
            "timestamp":   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error":       error_msg,
        }
        _test_results.append(res)
        import json
        try:
            with open("worker_results.jsonl", "a", encoding="utf-8") as wf:
                wf.write(json.dumps(res) + "\n")
        except Exception: pass


@pytest.hookimpl(tryfirst=True)
def pytest_runtest_teardown(item, nextitem):
    if "driver" in item.fixturenames or "fresh_driver" in item.fixturenames:
        time.sleep(1.0)
    else:
        time.sleep(0.1)


# ─── Excel report ─────────────────────────────────────────────────────────────
def pytest_sessionfinish(session, exitstatus):
    import json
    import os
    global _test_results
    _test_results = []
    if os.path.exists("worker_results.jsonl"):
        with open("worker_results.jsonl", "r", encoding="utf-8") as f:
            for line in f:
                if line.strip(): _test_results.append(json.loads(line))
    if not _test_results:
        return
    now  = datetime.datetime.now()
    name = f"E2E_Test_Report_LegalRiskAnalyzer_{now.strftime('%Y-%m-%dT%H-%M-%S')}.xlsx"
    path = os.path.join(os.path.dirname(__file__), name)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    NAVY = "1B3A6B"; ACCENT = "00B4D8"; WHITE = "FFFFFF"
    LGRAY = "F0F4F8"; ALT = "E8F4FD"

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(bold=False, color="000000", sz=11):
        return Font(bold=bold, color=color, size=sz, name="Calibri")
    def aln(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    def bdr():
        s = Side(style="thin", color="B0BEC5")
        return Border(left=s, right=s, top=s, bottom=s)

    H_FILL = fill(NAVY); H_FONT = font(True, WHITE, 11)
    C_AL = aln("center"); L_AL = aln("left", True)
    P_FILL = fill("C8E6C9"); P_FONT = font(True, "1B5E20")
    F_FILL = fill("FFCDD2"); F_FONT = font(True, "B71C1C")
    S_FILL = fill("FFF9C4"); S_FONT = font(True, "F57F17")

    def hrow(ws, r, cols, widths=None):
        for ci, v in enumerate(cols, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = H_FILL; c.font = H_FONT
            c.alignment = C_AL; c.border = bdr()
        if widths:
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[r].height = 24

    def ocell(ws, r, ci, outcome):
        c = ws.cell(row=r, column=ci, value=outcome.upper())
        c.border = bdr(); c.alignment = C_AL
        if outcome == "passed":   c.fill, c.font = P_FILL, P_FONT
        elif outcome == "failed": c.fill, c.font = F_FILL, F_FONT
        else:                     c.fill, c.font = S_FILL, S_FONT

    total   = len(_test_results)
    passed  = sum(1 for r in _test_results if r["outcome"] == "passed")
    failed  = sum(1 for r in _test_results if r["outcome"] == "failed")
    skipped = sum(1 for r in _test_results if r["outcome"] == "skipped")
    pr      = f"{passed / total * 100:.1f}%" if total > 0 else "0.0%"
    dur     = round(sum(r["duration"] for r in _test_results), 1)
    cats    = sorted(set(r["category"] for r in _test_results))

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:G1")
    t = ws1["A1"]
    t.value = "Legal Risk Analyzer — End-to-End Test Report"
    t.fill = fill(NAVY); t.font = font(True, ACCENT, 16); t.alignment = aln("center")
    ws1.row_dimensions[1].height = 44
    ws1.merge_cells("A2:G2")
    s = ws1["A2"]
    s.value = f"Generated: {now.strftime('%Y-%m-%d %H:%M:%S')}   |   Environment: Production"
    s.fill = fill("2D3268"); s.font = font(False, "94A3B8", 10); s.alignment = aln("center")
    ws1.row_dimensions[2].height = 20

    hrow(ws1, 4, ["Total Tests","Passed","Failed","Skipped","Pass Rate","Duration (s)"],
         [16,14,14,14,14,16])
    for ci, v in enumerate([total, passed, failed, skipped, pr, dur], 1):
        c = ws1.cell(row=5, column=ci, value=v)
        c.border = bdr(); c.alignment = C_AL
        c.fill = fill(LGRAY); c.font = font(True, NAVY, 13)
    ws1.row_dimensions[5].height = 32

    ws1["I4"] = "Outcome"; ws1["J4"] = "Count"
    ws1["I5"] = "Passed";  ws1["J5"] = passed
    ws1["I6"] = "Failed";  ws1["J6"] = failed
    ws1["I7"] = "Skipped"; ws1["J7"] = skipped
    if total > 0:
        pie = PieChart()
        pie.add_data(Reference(ws1, min_col=10, min_row=4, max_row=7), titles_from_data=True)
        pie.set_categories(Reference(ws1, min_col=9, min_row=5, max_row=7))
        pie.title = "Test Results Distribution"
        pie.width = 14; pie.height = 10
        dp0 = DataPoint(idx=0); dp0.graphicalProperties.solidFill = "4CAF50"
        dp1 = DataPoint(idx=1); dp1.graphicalProperties.solidFill = "F44336"
        dp2 = DataPoint(idx=2); dp2.graphicalProperties.solidFill = "FFC107"
        pie.series[0].data_points = [dp0, dp1, dp2]
        ws1.add_chart(pie, "A7")

    cat_start = 22
    hrow(ws1, cat_start,
         ["Category / File","Total","Passed","Failed","Skipped","Pass Rate %","Test Type"],
         [42,10,10,10,10,14,20])
    for i, cat in enumerate(cats, 1):
        cr = [r for r in _test_results if r["category"] == cat]
        ct = len(cr); cp = sum(1 for r in cr if r["outcome"]=="passed")
        cf = sum(1 for r in cr if r["outcome"]=="failed")
        cs = sum(1 for r in cr if r["outcome"]=="skipped")
        cpr = f"{cp/ct*100:.1f}%" if ct > 0 else "0.0%"
        alt = i % 2 == 0
        for ci, v in enumerate([cat,ct,cp,cf,cs,cpr,_classify_type(cat)], 1):
            c = ws1.cell(row=cat_start+i, column=ci, value=v)
            c.border = bdr(); c.alignment = L_AL if ci==1 else C_AL
            if alt: c.fill = fill(ALT)
        ws1.row_dimensions[cat_start+i].height = 20

    # ── Sheet 2: Execution Log ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Execution Log")
    ws2.sheet_view.showGridLines = False
    hrow(ws2, 1, ["#","TC ID","Category","Test Function","Description","Type",
                  "Outcome","Duration (s)","Timestamp"],
         [5,10,34,48,58,16,12,14,22])
    for ri, r in enumerate(_test_results, 2):
        alt = ri % 2 == 0
        vals = [ri-1, r["tc_id"], r["category"], r["test_name"],
                r["description"], r["type"], None, round(r["duration"],2), r["timestamp"]]
        for ci, v in enumerate(vals, 1):
            if ci == 7:
                ocell(ws2, ri, ci, r["outcome"])
            else:
                c = ws2.cell(row=ri, column=ci, value=v)
                c.border = bdr()
                c.alignment = L_AL if ci in (3,4,5) else C_AL
                if alt: c.fill = fill(ALT)
        ws2.row_dimensions[ri].height = 20

    # ── Sheet 3: Passed ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Passed Tests")
    ws3.sheet_view.showGridLines = False
    hrow(ws3, 1, ["#","TC ID","Category","Test Function","Description","Type",
                  "Duration (s)","Timestamp"], [5,10,34,48,58,16,14,22])
    pr_res = [r for r in _test_results if r["outcome"]=="passed"]
    for ri, r in enumerate(pr_res, 2):
        alt = ri % 2 == 0
        for ci, v in enumerate([ri-1, r["tc_id"], r["category"], r["test_name"],
                                  r["description"], r["type"],
                                  round(r["duration"],2), r["timestamp"]], 1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.border = bdr()
            c.fill = fill("DCEDC8") if alt else P_FILL
            c.alignment = L_AL if ci in (3,4,5) else C_AL
        ws3.row_dimensions[ri].height = 20

    # ── Sheet 4: Failed ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Failed Tests")
    ws4.sheet_view.showGridLines = False
    hrow(ws4, 1, ["#","TC ID","Category","Test Function","Description","Type",
                  "Duration (s)","Timestamp","Error / Reason"],
         [5,10,34,48,55,16,14,22,80])
    fl_res = [r for r in _test_results if r["outcome"]=="failed"]
    for ri, r in enumerate(fl_res, 2):
        for ci, v in enumerate([ri-1, r["tc_id"], r["category"], r["test_name"],
                                  r["description"], r["type"],
                                  round(r["duration"],2), r["timestamp"],
                                  r["error"][:5000]], 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.border = bdr(); c.fill = F_FILL
            c.alignment = L_AL if ci in (4,5,9) else C_AL
        ws4.row_dimensions[ri].height = 22

    # ── Sheet 5: Skipped ──────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Skipped Tests")
    ws5.sheet_view.showGridLines = False
    hrow(ws5, 1, ["#","TC ID","Category","Test Function","Description","Type","Timestamp"],
         [5,10,34,48,58,16,22])
    sk_res = [r for r in _test_results if r["outcome"]=="skipped"]
    for ri, r in enumerate(sk_res, 2):
        for ci, v in enumerate([ri-1, r["tc_id"], r["category"], r["test_name"],
                                  r["description"], r["type"], r["timestamp"]], 1):
            c = ws5.cell(row=ri, column=ci, value=v)
            c.border = bdr(); c.fill = S_FILL
            c.alignment = L_AL if ci in (4,5) else C_AL
        ws5.row_dimensions[ri].height = 20

    # ── Sheet 6: Category Analysis ────────────────────────────────────────────
    ws6 = wb.create_sheet("Category Analysis")
    ws6.sheet_view.showGridLines = False
    hrow(ws6, 1, ["Category","Total","Passed","Failed","Skipped","Pass Rate %"],
         [42,10,10,10,10,14])
    for i, cat in enumerate(cats, 2):
        cr = [r for r in _test_results if r["category"]==cat]
        ct = len(cr); cp = sum(1 for r in cr if r["outcome"]=="passed")
        cf = sum(1 for r in cr if r["outcome"]=="failed")
        cs = sum(1 for r in cr if r["outcome"]=="skipped")
        cpr = round(cp/ct*100, 1) if ct > 0 else 0.0
        alt = i % 2 == 0
        for ci, v in enumerate([cat,ct,cp,cf,cs,cpr], 1):
            c = ws6.cell(row=i, column=ci, value=v)
            c.border = bdr(); c.alignment = L_AL if ci==1 else C_AL
            if alt: c.fill = fill(ALT)
        ws6.row_dimensions[i].height = 20

    if cats:
        bar = BarChart()
        bar.type = "bar"; bar.grouping = "stacked"
        bar.title = "Test Results by Category"
        bar.y_axis.title = "Count"; bar.x_axis.title = "Category"
        bar.width = 28; bar.height = 14
        nr = 1 + len(cats)
        bar.add_data(Reference(ws6,min_col=3,min_row=1,max_row=nr), titles_from_data=True)
        bar.add_data(Reference(ws6,min_col=4,min_row=1,max_row=nr), titles_from_data=True)
        bar.add_data(Reference(ws6,min_col=5,min_row=1,max_row=nr), titles_from_data=True)
        bar.set_categories(Reference(ws6,min_col=1,min_row=2,max_row=nr))
        bar.series[0].graphicalProperties.solidFill = "4CAF50"
        bar.series[1].graphicalProperties.solidFill = "F44336"
        bar.series[2].graphicalProperties.solidFill = "FFC107"
        ws6.add_chart(bar, "H2")

    wb.save(path)
    print(f"\n[+] EXCEL REPORT: {path}")
    print(f"    Total={total} | Passed={passed} | Failed={failed} | Skipped={skipped} | PassRate={pr}\n")
