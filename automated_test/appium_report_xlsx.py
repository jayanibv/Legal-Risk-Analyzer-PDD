import json
import os
import datetime
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

REPORT_JSON = os.path.join(os.path.dirname(__file__), "appium_report.json")
NOW = datetime.datetime.utcnow()
OUTFILE = os.path.join(
    os.path.dirname(__file__),
    f"Appium_Mobile_Test_Report_LegalRiskAnalyzer_{NOW.strftime('%Y-%m-%dT%H-%M-%S')}.xlsx"
)

NAVY = "1B1F3B"
ACCENT = "00E5FF"
WHITE = "FFFFFF"

def _fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def build():
    if not os.path.exists(REPORT_JSON):
        print(f"[!] {REPORT_JSON} not found. Run appium_runner.py first.")
        return

    with open(REPORT_JSON, encoding="utf-8") as f:
        records = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appium Test Results"
    ws.sheet_view.showGridLines = False

    # Header
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Legal Risk Analyzer — Mobile Appium E2E Report"
    t.fill = _fill(NAVY)
    t.font = _font(True, ACCENT, 15)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 38

    # Table Headers
    headers = ["#", "Test Case", "Status", "Error Message / Notes"]
    widths = [5, 40, 15, 60]
    for ci, val in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=val)
        c.fill = _fill(NAVY); c.font = _font(True, WHITE); c.alignment = _align("center"); c.border = _border()
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci-1]

    # Data
    passed = sum(1 for r in records if r.get("status") == "PASS")
    total = len(records)

    for ri, r in enumerate(records, 4):
        status = r.get("status", "FAIL")
        is_pass = status == "PASS"
        
        # ID
        c_id = ws.cell(row=ri, column=1, value=ri-3)
        c_id.border = _border()
        
        # Test Case
        c_tc = ws.cell(row=ri, column=2, value=r.get("test_case", ""))
        c_tc.border = _border()
        
        # Status
        c_st = ws.cell(row=ri, column=3, value=status)
        c_st.border = _border()
        c_st.fill = _fill("90EE90" if is_pass else "FF6B6B")
        c_st.font = _font(True, "1A2E00" if is_pass else "4B0000")
        c_st.alignment = _align("center")
        
        # Error
        c_err = ws.cell(row=ri, column=4, value=r.get("error", ""))
        c_err.border = _border()
        c_err.alignment = _align("left", wrap=True)
        ws.row_dimensions[ri].height = 25 if not r.get("error") else 40

    wb.save(OUTFILE)
    print(f"[OK] Appium Excel report saved: {OUTFILE}")
    print(f"     Total Tests: {total} | Passed: {passed} | Failed: {total - passed}")

if __name__ == "__main__":
    build()
