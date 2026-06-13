"""
dast_report_xlsx.py — Convert report.json to DAST_Security_Report.xlsx
Matches the style of the reference E2E report.
Usage: python automated_test/dast_report_xlsx.py
"""
import json, os, datetime, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

REPORT_JSON  = os.path.join(os.path.dirname(__file__), "report.json")
BASE_URL_CFG = os.path.join(os.path.dirname(__file__), "input.json")
NOW          = datetime.datetime.utcnow()
OUTFILE      = os.path.join(
    os.path.dirname(__file__),
    f"DAST_Security_Report_LegalRiskAnalyzer_{NOW.strftime('%Y-%m-%dT%H-%M-%S')}.xlsx"
)

SEV_COLORS = {
    "CRITICAL": ("4B0000", "FF6B6B"),
    "HIGH":     ("451A1A", "FF9966"),
    "MEDIUM":   ("3D2800", "FFC107"),
    "LOW":      ("1A2E00", "90EE90"),
    "INFO":     ("1A1A2E", "8B9FD4"),
}
NAVY   = "1B1F3B"
ACCENT = "00E5FF"
WHITE  = "FFFFFF"

def _fill(hex_c):    return PatternFill("solid", fgColor=hex_c)
def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, cols, widths=None):
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = _fill(NAVY); c.font = _font(True, WHITE); c.alignment = _align("center"); c.border = _border()
    if widths:
        from openpyxl.utils import get_column_letter
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

def data_row(ws, row, vals, sev="INFO", alt=False):
    bg, fg = SEV_COLORS.get(sev, ("FFFFFF","000000"))
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.border = _border()
        if ci == len(vals):   # severity column
            c.fill = _fill(bg); c.font = _font(True, fg); c.alignment = _align("center")
        else:
            c.fill = _fill("F8FAFF" if alt else "FFFFFF")
            c.alignment = _align("left", wrap=(ci > 3))

def build():
    if not os.path.exists(REPORT_JSON):
        print(f"[!] {REPORT_JSON} not found. Run dast_runner.py first.")
        return

    with open(REPORT_JSON, encoding="utf-8") as f:
        records = json.load(f)

    with open(BASE_URL_CFG) as f:
        base_url = json.load(f).get("baseUrl", "N/A")

    findings = [r for r in records if r.get("finding")]
    total    = len(records)
    passed   = total - len(findings)

    sev_counts = {}
    for r in findings:
        s = r.get("severity", "INFO")
        sev_counts[s] = sev_counts.get(s, 0) + 1

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "Legal Risk Analyzer — DAST Security Assessment Report"
    t.fill = _fill(NAVY); t.font = _font(True, ACCENT, 15); t.alignment = _align("center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = f"Generated: {NOW.strftime('%Y-%m-%d %H:%M:%S UTC')}  |  Target: {base_url}"
    s.fill = _fill("2D3268"); s.font = _font(False, "94A3B8", 10); s.alignment = _align("center")
    ws.row_dimensions[2].height = 20

    header_row(ws, 4, ["Target URL", "Total Probes", "Findings", "CRITICAL", "HIGH", "MEDIUM", "LOW"],
               [48, 14, 12, 12, 12, 12, 12])
    data_row(ws, 5,
             [base_url, total, len(findings),
              sev_counts.get("CRITICAL",0), sev_counts.get("HIGH",0),
              sev_counts.get("MEDIUM",0), sev_counts.get("LOW",0)], "INFO")

    # Category breakdown
    ws["A7"] = "Category Breakdown"
    ws["A7"].fill = _fill(NAVY); ws["A7"].font = _font(True, ACCENT, 12)
    ws.merge_cells("A7:G7")
    header_row(ws, 8, ["Category", "Probes", "Findings", "Critical", "High", "Medium", "Info"])
    cats = {}
    for r in records:
        c = r.get("test_category","?")
        cats.setdefault(c, {"probes":0,"findings":0,"CRITICAL":0,"HIGH":0,"MEDIUM":0,"INFO":0})
        cats[c]["probes"] += 1
        if r.get("finding"):
            cats[c]["findings"] += 1
            cats[c][r.get("severity","INFO")] = cats[c].get(r.get("severity","INFO"),0)+1
    for ri, (cat, st) in enumerate(cats.items(), 9):
        data_row(ws, ri, [cat, st["probes"], st["findings"],
                          st.get("CRITICAL",0), st.get("HIGH",0), st.get("MEDIUM",0), st.get("INFO",0)],
                 "HIGH" if st["findings"] else "INFO", ri%2==0)

    # ── Sheet 2: Findings ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Findings")
    ws2.sheet_view.showGridLines = False
    header_row(ws2, 1, ["#","Endpoint","Method","Role/Context","Status","Expected",
                         "Category","Note","Severity"],
               [5, 30, 8, 18, 9, 9, 22, 62, 12])
    for ri, r in enumerate([x for x in records if x.get("finding")], 2):
        sev = r.get("severity","INFO")
        data_row(ws2, ri,
                 [ri-1, r.get("endpoint",""), r.get("method",""),
                  str(r.get("role",""))[:30], str(r.get("status","")),
                  str(r.get("expected_status","")),
                  r.get("test_category",""), r.get("note","")[:200], sev], sev, ri%2==0)
        ws2.row_dimensions[ri].height = 40

    # ── Sheet 3: All Tests ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("All Tests")
    ws3.sheet_view.showGridLines = False
    header_row(ws3, 1, ["#","Endpoint","Method","Role","Status","Expected",
                         "Finding","Category","Severity","Latency(ms)","Note","Timestamp"],
               [5, 30, 8, 18, 9, 9, 9, 22, 10, 11, 50, 22])
    for ri, r in enumerate(records, 2):
        sev = r.get("severity","INFO")
        f   = r.get("finding", False)
        data_row(ws3, ri,
                 [ri-1, r.get("endpoint",""), r.get("method",""),
                  str(r.get("role",""))[:30], str(r.get("status","")),
                  str(r.get("expected_status","")),
                  "YES" if f else "no",
                  r.get("test_category",""), sev,
                  r.get("response_time_ms",0),
                  r.get("note","")[:120], r.get("timestamp","")], sev if f else "INFO", ri%2==0)
        ws3.row_dimensions[ri].height = 22

    wb.save(OUTFILE)
    print(f"[OK] DAST Excel report saved: {OUTFILE}")
    print(f"     {total} probes | {len(findings)} findings | "
          f"CRIT={sev_counts.get('CRITICAL',0)} HIGH={sev_counts.get('HIGH',0)} "
          f"MED={sev_counts.get('MEDIUM',0)}")

if __name__ == "__main__":
    if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8","utf8"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    build()
