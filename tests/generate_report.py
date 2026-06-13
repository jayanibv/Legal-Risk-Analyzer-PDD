import json
import os
import datetime
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

TESTS_DIR = os.path.dirname(__file__)
REPORT_JSON = os.path.join(TESTS_DIR, ".results.json")
NOW = datetime.datetime.utcnow()
OUTFILE = os.path.join(
    TESTS_DIR,
    f"E2E_Test_Report_LegalRiskAnalyzer_{NOW.strftime('%Y-%m-%dT%H-%M-%S')}.xlsx"
)

NAVY = "1B1F3B"
ACCENT = "00E5FF"
WHITE = "FFFFFF"

def _fill(hex_c):    return PatternFill("solid", fgColor=hex_c)
def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, cols, widths=None):
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = _fill(NAVY); c.font = _font(True, WHITE); c.alignment = _align("center"); c.border = _border()
    if widths:
        from openpyxl.utils import get_column_letter
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

def data_row(ws, row, vals, outcome="passed", alt=False):
    fg = "000000"
    if outcome == "failed":
        bg = "FF6B6B"
        fg = "FFFFFF"
    elif outcome == "passed":
        bg = "90EE90"
    else:
        bg = "FFC107" # error or skipped
    
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.border = _border()
        if ci == len(vals):   # outcome column
            c.fill = _fill(bg); c.font = _font(True, fg); c.alignment = _align("center")
        else:
            c.fill = _fill("F8FAFF" if alt else "FFFFFF")
            c.alignment = _align("left", wrap=(ci > 3))

def build():
    if "--demo" in sys.argv:
        print("[DEMO] Generating demo report.")
        tests = [{"nodeid": "demo/test_demo.py::test_1", "outcome": "passed", "duration": 0.1, "setup": {"duration": 0.0}}]
        summary = {"passed": 1, "total": 1, "duration": 0.1}
    else:
        if not os.path.exists(REPORT_JSON):
            print(f"[!] {REPORT_JSON} not found. Run tests first.")
            return
        with open(REPORT_JSON, encoding="utf-8") as f:
            data = json.load(f)
        tests = data.get("tests", [])
        summary = data.get("summary", {})

    total = summary.get("total", len(tests))
    passed = summary.get("passed", sum(1 for t in tests if t.get("outcome") == "passed"))
    failed = summary.get("failed", sum(1 for t in tests if t.get("outcome") == "failed"))
    errors = summary.get("error", 0)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "Legal Risk Analyzer — E2E Test Report"
    t.fill = _fill(NAVY); t.font = _font(True, ACCENT, 15); t.alignment = _align("center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value = f"Generated: {NOW.strftime('%Y-%m-%d %H:%M:%S UTC')}"
    s.fill = _fill("2D3268"); s.font = _font(False, "94A3B8", 10); s.alignment = _align("center")
    ws.row_dimensions[2].height = 20

    header_row(ws, 4, ["Total Tests", "Passed", "Failed", "Errors", "Duration (s)"], [15, 15, 15, 15, 15])
    
    ws.cell(row=5, column=1, value=total).border = _border()
    ws.cell(row=5, column=2, value=passed).border = _border()
    ws.cell(row=5, column=3, value=failed).border = _border()
    ws.cell(row=5, column=4, value=errors).border = _border()
    ws.cell(row=5, column=5, value=round(summary.get("duration", 0), 2)).border = _border()
    
    for ci in range(1, 6):
        ws.cell(row=5, column=ci).alignment = _align("center")
        ws.cell(row=5, column=ci).fill = _fill("FFFFFF")

    # ── Sheet 2: All Tests ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Tests")
    ws2.sheet_view.showGridLines = False
    header_row(ws2, 1, ["#", "File", "Test Name", "Duration (s)", "Outcome"], [5, 40, 60, 15, 15])
    
    for ri, r in enumerate(tests, 2):
        nodeid = r.get("nodeid", "")
        parts = nodeid.split("::")
        file = parts[0]
        name = parts[-1] if len(parts) > 1 else ""
        outcome = r.get("outcome", "unknown")
        dur = r.get("setup", {}).get("duration", 0) + r.get("call", {}).get("duration", 0) + r.get("teardown", {}).get("duration", 0)
        
        data_row(ws2, ri, [ri-1, file, name, round(dur, 2), outcome], outcome, ri%2==0)
        ws2.row_dimensions[ri].height = 22

    wb.save(OUTFILE)
    print(f"[OK] E2E Excel report saved: {OUTFILE}")
    print(f"     {total} tests | {passed} passed | {failed} failed | {errors} errors")

if __name__ == "__main__":
    if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8","utf8"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    build()
