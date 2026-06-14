"""
conftest.py - Shared fixtures and configuration for Legal Risk Analyzer E2E tests
"""
import pytest
import time
import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# ─── Configuration ────────────────────────────────────────────────────────────
BASE_URL = "https://legal-risk-analyzer.up.railway.app"      # FastAPI backend
FRONTEND_URL = "https://legal-risk-analyzer-pdd.vercel.app"  # Expo/React web frontend

# Test credentials (use a dedicated test account)
TEST_EMAIL = "testuser_e2e@legalrisk.dev"
TEST_PASSWORD = "TestPass@123"
TEST_NAME = "E2E Test User"
TEST_DOB = "1995-06-15"
TEST_SECURITY = "testfriend"

# ─── Shared Driver Fixture ────────────────────────────────────────────────────
def _make_driver():
    """Factory: create a fresh Chrome WebDriver instance."""
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    # Uncomment for headless CI:
    options.add_argument("--headless=new")
    service = Service(ChromeDriverManager().install())
    drv = webdriver.Chrome(service=service, options=options)
    drv.set_page_load_timeout(30)
    drv.implicitly_wait(8)
    return drv


def _quit_driver(drv):
    """Safely quit a driver, swallowing any InvalidSessionIdException."""
    try:
        drv.quit()
    except Exception:
        pass


@pytest.fixture(scope="class")
def driver():
    """
    Class-scoped Chrome WebDriver.
    One browser instance per test class — isolates crashes between classes.
    """
    drv = _make_driver()
    yield drv
    _quit_driver(drv)


@pytest.fixture(scope="module")
def fresh_driver():
    """
    Module-scoped Chrome WebDriver — a fresh browser per test module.
    Useful when isolation between test groups is needed.
    """
    drv = _make_driver()
    yield drv
    _quit_driver(drv)


# ─── Helpers ──────────────────────────────────────────────────────────────────
def get_timestamp():
    return datetime.datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S")

# ─── Excel Reporting Engine ───────────────────────────────────────────────────
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

# Global list to store results
_test_results = []

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    
    # We only care about the actual test call, or setup/teardown if they fail
    if rep.when == "call" or (rep.when in ["setup", "teardown"] and rep.failed):
        error_msg = ""
        if rep.failed:
            error_msg = rep.longreprtext if hasattr(rep, 'longreprtext') else str(rep.longrepr)
            
        category = item.nodeid.split("::")[0].split("/")[-1].split("\\")[-1].replace(".py", "")
        test_name = item.name
        
        _test_results.append({
            "category": category,
            "test_name": test_name,
            "outcome": rep.outcome,
            "duration": getattr(rep, "duration", 0),
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error": error_msg
        })

@pytest.hookimpl(tryfirst=True)
def pytest_runtest_teardown(item, nextitem):
    """Throttle requests to prevent overwhelming the server"""
    time.sleep(1.5)

def pytest_sessionfinish(session, exitstatus):
    now = datetime.datetime.now()
    report_filename = f"E2E_Test_Report_LegalRiskAnalyzer_{now.strftime('%Y-%m-%dT%H-%M-%S')}.xlsx"
    tests_dir = os.path.dirname(__file__)
    filepath = os.path.join(tests_dir, report_filename)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    left_aligned_text = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pass_font = Font(color="006100")
    
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(color="9C0006")
    
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    skip_font = Font(color="9C6500")

    ws_summary = wb.create_sheet("Summary")
    ws_passed = wb.create_sheet("Passed Tests")
    ws_failed = wb.create_sheet("Failed Tests")
    ws_log = wb.create_sheet("Execution Log")

    total = len(_test_results)
    passed = sum(1 for r in _test_results if r["outcome"] == "passed")
    failed = sum(1 for r in _test_results if r["outcome"] == "failed")
    skipped = sum(1 for r in _test_results if r["outcome"] == "skipped")

    def apply_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned_text
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # --- Summary Tab ---
    apply_header(ws_summary, ["Metric", "Value"])
    metrics = [
        ("Total Tests", total, None, None),
        ("Passed", passed, pass_fill, pass_font),
        ("Failed", failed, fail_fill, fail_font),
        ("Skipped", skipped, skip_fill, skip_font)
    ]
    
    for row, (metric, val, fill, font) in enumerate(metrics, 2):
        c1 = ws_summary.cell(row=row, column=1, value=metric)
        c2 = ws_summary.cell(row=row, column=2, value=val)
        c1.border = thin_border
        c2.border = thin_border
        if fill:
            c1.fill = fill
            c2.fill = fill
        if font:
            c1.font = font
            c2.font = font
            
    # Pie Chart
    if total > 0:
        pie = PieChart()
        labels = Reference(ws_summary, min_col=1, min_row=3, max_row=5)
        data = Reference(ws_summary, min_col=2, min_row=2, max_row=5) # Includes title for series name
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Test Execution Results"
        
        # Color the slices: Passed (Green), Failed (Red), Skipped (Orange)
        slice_passed = DataPoint(idx=0)
        slice_passed.graphicalProperties.solidFill = "00B050"
        slice_failed = DataPoint(idx=1)
        slice_failed.graphicalProperties.solidFill = "FF0000"
        slice_skipped = DataPoint(idx=2)
        slice_skipped.graphicalProperties.solidFill = "FFC000"
        
        pie.series[0].data_points = [slice_passed, slice_failed, slice_skipped]
        ws_summary.add_chart(pie, "D2")

    # Category Breakdown
    cat_start_row = 8
    headers = ["Category", "Total", "Passed", "Failed", "Skipped", "Pass Rate %"]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=cat_start_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned_text
        cell.border = thin_border

    categories = set(r["category"] for r in _test_results)
    for i, cat in enumerate(sorted(categories), 1):
        cat_results = [r for r in _test_results if r["category"] == cat]
        cat_total = len(cat_results)
        cat_passed = sum(1 for r in cat_results if r["outcome"] == "passed")
        cat_failed = sum(1 for r in cat_results if r["outcome"] == "failed")
        cat_skipped = sum(1 for r in cat_results if r["outcome"] == "skipped")
        pass_rate = f"{(cat_passed/cat_total)*100:.1f}%" if cat_total > 0 else "0.0%"
        
        row_data = [cat, cat_total, cat_passed, cat_failed, cat_skipped, pass_rate]
        for col, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=cat_start_row + i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_aligned_text if col > 1 else left_aligned_text

    # --- Passed Tests Tab ---
    apply_header(ws_passed, ["Category", "Test Name", "Duration (s)", "Timestamp"])
    passed_results = [r for r in _test_results if r["outcome"] == "passed"]
    for row, res in enumerate(passed_results, 2):
        row_data = [res["category"], res["test_name"], round(res["duration"], 2), res["timestamp"]]
        for col, val in enumerate(row_data, 1):
            cell = ws_passed.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = left_aligned_text

    # --- Failed Tests Tab ---
    apply_header(ws_failed, ["Category", "Test Name", "Error", "Duration (s)", "Timestamp"])
    failed_results = [r for r in _test_results if r["outcome"] == "failed"]
    for row, res in enumerate(failed_results, 2):
        row_data = [res["category"], res["test_name"], res["error"][:30000], round(res["duration"], 2), res["timestamp"]]
        for col, val in enumerate(row_data, 1):
            cell = ws_failed.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = left_aligned_text

    # --- Execution Log Tab ---
    apply_header(ws_log, ["Category", "Test Name", "Outcome", "Duration (s)", "Timestamp"])
    for row, res in enumerate(_test_results, 2):
        row_data = [res["category"], res["test_name"], res["outcome"], round(res["duration"], 2), res["timestamp"]]
        for col, val in enumerate(row_data, 1):
            cell = ws_log.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = left_aligned_text
            if col == 3: # Outcome column
                if res["outcome"] == "passed":
                    cell.fill, cell.font = pass_fill, pass_font
                elif res["outcome"] == "failed":
                    cell.fill, cell.font = fail_fill, fail_font
                else:
                    cell.fill, cell.font = skip_fill, skip_font

    wb.save(filepath)
    print(f"\n[+] EXCEL REPORT GENERATED: {filepath}\n")
