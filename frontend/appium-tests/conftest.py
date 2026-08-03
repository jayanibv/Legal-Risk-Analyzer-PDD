"""
conftest.py
===========
Shared Appium fixtures, capabilities, and Excel report generation
for the Legal Risk Analyzer mobile E2E test suite.

APP FLOW (as discovered from source code):
  App Launch
    └─► index.tsx
          ├─ authenticated? ──► /(drawer)  [Home Dashboard]
          └─ not auth?     ──► /onboarding [3 slides]
                                  ├─ Skip ──────────────────► /login
                                  └─ Next/Next/Get Started ──► /signup

Each test class uses the 'driver' fixture (scope=class) which:
  1. Launches the app with fullReset=True (always starts fresh at onboarding)
  2. Calls skip_onboarding() to get to login immediately
  3. Then tests do whatever navigation they need from login
"""

import pytest
import time
import os
import uuid
from datetime import datetime
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.common.options import ArgOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# ─── CONFIG ──────────────────────────────────────────────────────────────────
APK_PATH       = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "LegalRiskAnalyzer.apk")
)
APP_PACKAGE    = "com.jayani.legalriskanalyzer"
APP_ACTIVITY   = ".MainActivity"
WAIT_TIMEOUT   = 20   # seconds for explicit waits
SHORT_WAIT     = 5

# Test account credentials
TEST_EMAIL     = "appiumtest@legalrisk.dev"
TEST_PASSWORD  = "AppiumTest@999"
TEST_NAME      = "Appium Tester"
TEST_DOB       = "1995-06-15"
TEST_SECURITY  = "appiumfriend"

# Onboarding text signatures (from onboarding.tsx)
ONBOARDING_TITLES = [
    "Analyze Contracts in Seconds",
    "AI-Powered Risk Detection",
    "Bank-Grade Security",
]
ONBOARDING_BUTTONS = ["Next", "Get Started", "Skip"]

# ─── BROWSERSTACK / LOCAL MODE ────────────────────────────────────────────────
# When BROWSERSTACK_USERNAME env var is set (i.e. running in GitHub Actions),
# tests run on BrowserStack real devices — no emulator needed.
# When NOT set (i.e. running locally), tests use your local Appium + emulator.
BS_USERNAME  = os.environ.get("BROWSERSTACK_USERNAME", "")
BS_ACCESSKEY = os.environ.get("BROWSERSTACK_ACCESS_KEY", "")
BS_APP_URL   = os.environ.get("BROWSERSTACK_APP_URL", "")   # uploaded APK url
USE_BROWSERSTACK = bool(BS_USERNAME and BS_ACCESSKEY and BS_APP_URL)

if USE_BROWSERSTACK:
    # BrowserStack remote hub — no local Appium server needed
    APPIUM_HOST = f"https://{BS_USERNAME}:{BS_ACCESSKEY}@hub.browserstack.com/wd/hub"
else:
    # Local Appium server (dev / emulator)
    APPIUM_HOST = "http://127.0.0.1:4723"

# Global results collector
_results = []


# ─── CAPABILITY BUILDER ───────────────────────────────────────────────────────
def get_options(full_reset=False, no_reset=False):
    """
    Build Appium 2.x / Selenium 4.x compatible options.
    Automatically switches between BrowserStack (CI) and local emulator (dev).

    BrowserStack mode  → uses real cloud Android device, no emulator
    Local mode         → uses local Appium + emulator
    """
    opts = ArgOptions()

    if USE_BROWSERSTACK:
        # ── BrowserStack W3C capabilities ────────────────────────────────────
        opts.set_capability("platformName", "Android")
        opts.set_capability("bstack:options", {
            "userName":        BS_USERNAME,
            "accessKey":       BS_ACCESSKEY,
            "projectName":     "Legal Risk Analyzer",
            "buildName":       "Appium E2E Build",
            "sessionName":     "Android Tests",
            "deviceName":      "Samsung Galaxy S23",
            "osVersion":       "13.0",
            "autoGrantPermissions": True,
            "newCommandTimeout": 120,
            "debug":           True,
            "networkLogs":     True,
        })
        opts.set_capability("appium:app",         BS_APP_URL)   # bs://... URL
        opts.set_capability("appium:appPackage",  APP_PACKAGE)
        opts.set_capability("appium:appActivity", APP_ACTIVITY)
        opts.set_capability("appium:automationName", "UiAutomator2")
        opts.set_capability("appium:noReset",     no_reset)
    else:
        # ── Local emulator capabilities ───────────────────────────────────────
        opts.set_capability("platformName",                             "Android")
        opts.set_capability("appium:deviceName",                       "Android Emulator")
        opts.set_capability("appium:app",                              APK_PATH)
        opts.set_capability("appium:appPackage",                       APP_PACKAGE)
        opts.set_capability("appium:appActivity",                      APP_ACTIVITY)
        opts.set_capability("appium:automationName",                   "UiAutomator2")
        opts.set_capability("appium:noReset",                          no_reset)
        opts.set_capability("appium:fullReset",                        full_reset)
        opts.set_capability("appium:newCommandTimeout",                120)
        opts.set_capability("appium:autoGrantPermissions",             True)
        opts.set_capability("appium:uiautomator2ServerInstallTimeout", 60000)

    return opts


# ─── DRIVER FIXTURE ──────────────────────────────────────────────────────────
@pytest.fixture(scope="class")
def driver():
    """
    Create one Appium session per test class.
    BrowserStack: connects to real cloud device.
    Local: clears app data so onboarding shows, then skips to login.
    """
    opts = get_options(full_reset=False, no_reset=False)
    d = webdriver.Remote(command_executor=APPIUM_HOST, options=opts)
    d.implicitly_wait(WAIT_TIMEOUT)
    # Always skip onboarding first so tests start from login screen
    skip_onboarding(d)
    yield d
    d.quit()


@pytest.fixture(scope="function")
def driver_fresh():
    """
    Fresh session per onboarding test.
    BrowserStack: new cloud session per test.
    Local: clears app data only (no reinstall — avoids INSTALL_FAILED_PACKAGE_CHANGED).
    """
    opts = get_options(full_reset=False, no_reset=False)
    d = webdriver.Remote(command_executor=APPIUM_HOST, options=opts)
    d.implicitly_wait(WAIT_TIMEOUT)
    yield d
    d.quit()


# ─── ONBOARDING NAVIGATION ───────────────────────────────────────────────────
def is_on_onboarding(driver):
    """Return True if the current screen is the onboarding screen."""
    text = get_screen_text(driver)
    return any(title in text for title in ONBOARDING_TITLES) or "Skip" in text


def skip_onboarding(driver, timeout=20):
    """
    Skip the onboarding flow using the 'Skip' button.
    Loops until we actually see the login screen.
    """
    time.sleep(2)  # Give the app a moment to launch and settle
    deadline = time.time() + timeout
    while time.time() < deadline:
        text = fast_get_screen_text(driver).lower()
        
        # Already on login screen
        if any(kw in text for kw in ["sign in", "don't have", "forgot"]):
            return True
            
        # Already on home/dashboard
        if any(kw in text for kw in ["history", "upload", "chat", "settings"]):
            return True
            
        skip_btn = safe_find(driver, '//*[contains(@text, "Skip") or contains(@content-desc, "Skip")]', timeout=2)
        if skip_btn:
            try:
                force_tap(driver, skip_btn)
            except Exception:
                pass
            time.sleep(3)  # Wait for transition, loop will verify if it worked
        else:
            # Fallback: maybe we can't find Skip, try tapping Next instead to get through it
            btn = safe_find(driver, '//*[contains(@text, "Next") or contains(@text, "Get Started")]', timeout=1)
            if btn:
                try:
                    force_tap(driver, btn)
                except Exception:
                    pass
            time.sleep(1.5)
            
    return False


def navigate_through_onboarding(driver):
    """
    Go through all 3 onboarding slides using Next/Get Started.
    Ends on /signup screen.
    """
    time.sleep(2)
    deadline = time.time() + 20
    while time.time() < deadline:
        text = fast_get_screen_text(driver).lower()
        if any(kw in text for kw in ["sign up", "create account", "already have"]):
            return True
            
        btn = safe_find(driver, '//*[contains(@text, "Get Started") or contains(@text, "Next")]', timeout=2)
        if btn:
            try:
                force_tap(driver, btn)
            except Exception:
                pass
            time.sleep(1.5)
        else:
            time.sleep(1)
    return False


def navigate_to_login(driver, timeout=15):
    """
    Ensure driver is on the login screen.
    From any screen, get to /login.
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        text = fast_get_screen_text(driver).lower()
        
        # Already on login
        if any(kw in text for kw in ["sign in", "don't have", "forgot password"]):
            return True
            
        # On onboarding
        if any(title.lower() in text for title in ONBOARDING_TITLES) or "skip" in text:
            skip_onboarding(driver)
            continue
            
        # On signup - tap login link
        if any(kw in text for kw in ["sign up", "create account"]):
            login_link = safe_find(driver, '//*[contains(@text,"Sign In") or contains(@text,"Login") or contains(@text,"already")]', timeout=3)
            if login_link:
                try:
                    login_link.click()
                except Exception:
                    pass
                time.sleep(2)
                continue
                
        # Fallback - back button
        try:
            driver.back()
        except Exception:
            pass
        time.sleep(1.5)
        
    return False


# ─── HELPER FUNCTIONS ────────────────────────────────────────────────────────
def wait_for_element(driver, locator, by=AppiumBy.XPATH, timeout=WAIT_TIMEOUT):
    """Wait until an element is visible and return it."""
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, locator))
    )


def safe_find(driver, locator, by=AppiumBy.XPATH, timeout=SHORT_WAIT):
    """Return element or None without raising."""
    try:
        return WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, locator))
        )
    except (TimeoutException, NoSuchElementException):
        return None


def force_tap(driver, element):
    """Force tap an element using coordinates. Fixes React Native TouchableOpacity bugs."""
    try:
        location = element.location
        size = element.size
        x = int(location['x'] + (size['width'] / 2))
        y = int(location['y'] + (size['height'] / 2))
        driver.tap([(x, y)])
    except Exception:
        # Fallback if tap fails
        element.click()


def tap_element(driver, locator, by=AppiumBy.XPATH, timeout=WAIT_TIMEOUT):
    """Wait and tap an element."""
    el = wait_for_element(driver, locator, by, timeout)
    force_tap(driver, el)
    return el


def type_into(driver, locator, text, by=AppiumBy.XPATH, clear=True):
    """Find a field, clear it, then type text."""
    el = wait_for_element(driver, locator, by)
    if clear:
        el.clear()
    el.send_keys(text)
    return el


def scroll_down(driver, times=1):
    """Scroll down using UiAutomator2 scroll gesture."""
    for _ in range(times):
        size = driver.get_window_size()
        w, h = size["width"], size["height"]
        driver.swipe(w // 2, int(h * 0.75), w // 2, int(h * 0.25), 600)
        time.sleep(0.3)


def scroll_up(driver, times=1):
    """Scroll up."""
    for _ in range(times):
        size = driver.get_window_size()
        w, h = size["width"], size["height"]
        driver.swipe(w // 2, int(h * 0.25), w // 2, int(h * 0.75), 600)
        time.sleep(0.3)


def element_exists(driver, locator, by=AppiumBy.XPATH, timeout=SHORT_WAIT):
    """Return True if element exists within timeout."""
    return safe_find(driver, locator, by, timeout) is not None


def get_screen_text(driver):
    """Get all text on the current screen."""
    try:
        els = driver.find_elements(AppiumBy.XPATH, "//*[@text]")
        return " ".join(e.get_attribute("text") or "" for e in els)
    except Exception:
        return ""


def fast_get_screen_text(driver):
    """Get screen text without waiting. Prevents 10-second delays on blank transition screens."""
    driver.implicitly_wait(0)
    try:
        els = driver.find_elements(AppiumBy.XPATH, "//*[@text]")
        return " ".join(e.get_attribute("text") or "" for e in els)
    except Exception:
        return ""
    finally:
        driver.implicitly_wait(WAIT_TIMEOUT)


def navigate_back(driver, times=1):
    """Press back button N times."""
    for _ in range(times):
        driver.back()
        time.sleep(0.5)


def login_as_test_user(driver):
    """
    Navigate to login screen and log in with the test account.
    Handles onboarding if app is still on that screen.
    Returns True if login succeeded.
    """
    try:
        # Make sure we're on login screen first
        navigate_to_login(driver)
        time.sleep(1)

        # Find email field
        fields = driver.find_elements(AppiumBy.XPATH, '//android.widget.EditText')
        if len(fields) < 2:
            return False

        fields[0].clear()
        fields[0].send_keys(TEST_EMAIL)
        fields[1].clear()
        fields[1].send_keys(TEST_PASSWORD)

        # Tap login button
        btn = safe_find(driver, '//*[@text="Sign In" or @text="Login" or @text="Log In"]')
        if btn:
            btn.click()
            time.sleep(5)
            return True
        return False
    except Exception:
        return False


# ─── RESULT TRACKING ─────────────────────────────────────────────────────────
def record_result(tc_id, name, module, status, duration, notes=""):
    """Append a test result to the global results list."""
    _results.append({
        "tc_id":     tc_id,
        "name":      name,
        "module":    module,
        "status":    status,
        "duration":  round(duration, 3),
        "notes":     notes,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })


# ─── PYTEST HOOKS ────────────────────────────────────────────────────────────
@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report  = outcome.get_result()
    if report.when == "call":
        markers  = {m.name: m for m in item.iter_markers()}
        tc_id    = markers["tc"].args[0] if "tc" in markers else item.nodeid
        module   = item.module.__name__
        duration = report.duration
        status   = "PASS" if report.passed else ("FAIL" if report.failed else "SKIP")
        notes    = str(report.longreprtext) if report.failed else ""
        record_result(tc_id, item.name, module, status, duration, notes[:200])


# ─── EXCEL REPORT ────────────────────────────────────────────────────────────
def pytest_sessionfinish(session, exitstatus):
    if not _results or openpyxl is None:
        return

    ts       = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    out_path = os.path.join(os.path.dirname(__file__), f"Appium_Test_Report_{ts}.xlsx")
    wb       = openpyxl.Workbook()

    C_DARK   = "1A1A2E"
    C_NAVY   = "16213E"
    C_ACCENT = "0F3460"
    C_GREEN  = "27AE60"
    C_RED    = "E74C3C"
    C_GOLD   = "F39C12"
    C_WHITE  = "FFFFFF"
    C_LIGHT  = "ECF0F1"
    C_PASS   = "D5F5E3"
    C_FAIL   = "FADBD8"
    C_SKIP   = "FEF9E7"

    def fill(c):  return PatternFill("solid", fgColor=c)
    def bdr():
        s = Side(border_style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    total  = len(_results)
    passed = sum(1 for r in _results if r["status"] == "PASS")
    failed = sum(1 for r in _results if r["status"] == "FAIL")
    skip   = sum(1 for r in _results if r["status"] == "SKIP")
    prate  = round(passed / total * 100, 2) if total else 0

    # ── SHEET 1: SUMMARY ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 30

    ws1.merge_cells("A1:B1")
    c = ws1["A1"]
    c.value     = "Legal Risk Analyzer - Appium Mobile E2E Test Report"
    c.font      = Font(name="Segoe UI", bold=True, size=16, color=C_WHITE)
    c.fill      = fill(C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 42

    ws1.merge_cells("A2:B2")
    c = ws1["A2"]
    c.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  App: Legal Risk Analyzer Android"
    c.font      = Font(name="Segoe UI", size=10, italic=True, color="888888")
    c.fill      = fill(C_NAVY)
    c.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 20

    meta = [
        ("Platform",              "Android (Appium 2.x + UiAutomator2)"),
        ("App Package",           "com.jayani.legalriskanalyzer"),
        ("App Flow",              "Onboarding(3) → Login / Signup → Dashboard"),
        ("Test Framework",        "pytest + Appium-Python-Client v5"),
        ("Total Test Cases",      str(total)),
        ("Passed",                str(passed)),
        ("Failed",                str(failed)),
        ("Skipped",               str(skip)),
        ("Pass Rate",             f"{prate}%"),
    ]
    for idx, (k, v) in enumerate(meta, start=4):
        ws1.row_dimensions[idx].height = 24
        kc = ws1.cell(row=idx, column=1, value=k)
        vc = ws1.cell(row=idx, column=2, value=v)
        bg = C_LIGHT if idx % 2 == 0 else C_WHITE
        for c in (kc, vc):
            c.fill      = fill(bg)
            c.font      = Font(name="Segoe UI", bold=(c == kc), size=11)
            c.border    = bdr()
            c.alignment = Alignment(vertical="center", indent=1)
        if k == "Pass Rate":
            col = C_GREEN if prate >= 90 else (C_GOLD if prate >= 70 else C_RED)
            vc.font = Font(name="Segoe UI", bold=True, size=13, color=col)

    # Module breakdown
    modules = {}
    for r in _results:
        m = r["module"]
        if m not in modules:
            modules[m] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
        modules[m]["total"] += 1
        if r["status"] == "PASS":   modules[m]["pass"] += 1
        elif r["status"] == "FAIL": modules[m]["fail"] += 1
        else:                       modules[m]["skip"] += 1

    mod_start = len(meta) + 6
    ws1.merge_cells(f"A{mod_start}:F{mod_start}")
    c = ws1[f"A{mod_start}"]
    c.value     = "Module Breakdown"
    c.font      = Font(name="Segoe UI", bold=True, size=13, color=C_WHITE)
    c.fill      = fill(C_ACCENT)
    c.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[mod_start].height = 26

    for col_l, w in zip("BCDEF", [10, 10, 10, 10, 12]):
        ws1.column_dimensions[col_l].width = w

    h_row = mod_start + 1
    for col, hdr in enumerate(["Module", "Total", "Pass", "Fail", "Skip", "Rate %"], 1):
        c = ws1.cell(row=h_row, column=col, value=hdr)
        c.font      = Font(name="Segoe UI", bold=True, size=10, color=C_WHITE)
        c.fill      = fill(C_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()
    ws1.row_dimensions[h_row].height = 22

    for ridx, (mod, st) in enumerate(sorted(modules.items()), start=h_row + 1):
        rate = round(st["pass"] / st["total"] * 100, 1) if st["total"] else 0
        bg   = C_LIGHT if ridx % 2 == 0 else C_WHITE
        ws1.row_dimensions[ridx].height = 20
        for col, val in enumerate([mod.replace("test_0", "").replace("test_", "").replace("_", " ").title(),
                                    st["total"], st["pass"], st["fail"], st["skip"], f"{rate}%"], 1):
            c = ws1.cell(row=ridx, column=col, value=val)
            c.fill      = fill(bg)
            c.font      = Font(name="Segoe UI", size=10)
            c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                    vertical="center", indent=1 if col == 1 else 0)
            c.border    = bdr()
            if col == 6:
                rv = float(str(val).replace("%", ""))
                c.font = Font(name="Segoe UI", bold=True, size=10,
                              color=C_GREEN if rv >= 90 else (C_GOLD if rv >= 70 else C_RED))

    # ── SHEET 2: DETAILED RESULTS ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Detailed Results")
    ws2.sheet_view.showGridLines = False

    col_widths  = [10, 44, 30, 10, 12, 26, 22]
    col_headers = ["TC ID", "Test Name", "Module", "Status", "Duration(s)", "Notes", "Timestamp"]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.row_dimensions[1].height = 28
    for col, h in enumerate(col_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font      = Font(name="Segoe UI", bold=True, size=11, color=C_WHITE)
        c.fill      = fill(C_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()

    STATUS_BG = {"PASS": C_PASS, "FAIL": C_FAIL, "SKIP": C_SKIP}
    STATUS_FC = {"PASS": C_GREEN, "FAIL": C_RED, "SKIP": "E67E22"}

    for ridx, r in enumerate(_results, start=2):
        ws2.row_dimensions[ridx].height = 19
        bg   = STATUS_BG.get(r["status"], C_WHITE)
        vals = [r["tc_id"], r["name"], r["module"], r["status"],
                r["duration"], r["notes"], r["timestamp"]]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=ridx, column=col, value=val)
            c.fill      = fill(bg if col <= 4 else (C_LIGHT if ridx % 2 == 0 else C_WHITE))
            c.font      = Font(name="Segoe UI", size=10,
                               bold=(col == 4),
                               color=STATUS_FC.get(r["status"], "000000") if col == 4 else "000000")
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col in (1, 4, 5) else "left",
                                    wrap_text=(col == 6))
            c.border    = bdr()

    wb.save(out_path)
    print(f"\n[EXCEL] Appium report saved -> {out_path}")
    print(f"        Total={total} | Passed={passed} | Failed={failed} | Skipped={skip} | PassRate={prate}%")
